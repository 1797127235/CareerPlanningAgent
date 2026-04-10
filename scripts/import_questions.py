"""Import interview questions from xlsx files into InterviewQuestion table."""
from __future__ import annotations

import os
import re
import sys

# Add project root to path
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import openpyxl
from backend.db import engine, SessionLocal
from backend.db_models import Base, InterviewQuestion


def _detect_category(filename: str) -> tuple[str, str, str] | None:
    """Detect (skill_tag, question_category, node_id) from filename patterns."""
    lower = filename.lower()
    if "ai" in lower:
        return ("AI大模型", "technical", "AI工程师")
    if "mysql" in lower:
        return ("MySQL", "technical", "后端工程师")
    # Use byte patterns for Chinese chars that may appear garbled
    # Check original Chinese or garbled forms
    if re.search(r"并发|concurrent", filename, re.IGNORECASE):
        return ("Java并发", "technical", "Java后端工程师")
    if re.search(r"虚拟|jvm", filename, re.IGNORECASE):
        return ("JVM", "technical", "Java后端工程师")
    if re.search(r"集合|collection", filename, re.IGNORECASE):
        return ("Java集合", "technical", "Java后端工程师")
    # Fallback: check for java keyword
    if "java" in lower:
        # Try to distinguish by examining content later
        return ("Java基础", "technical", "Java后端工程师")
    return None


def assign_difficulty(idx: int, total: int) -> str:
    ratio = idx / total if total > 0 else 0
    if ratio < 0.4:
        return "easy"
    elif ratio < 0.75:
        return "medium"
    return "hard"


def import_from_dir(questions_dir: str, dry_run: bool = False):
    """Import all xlsx files from directory into DB."""
    Base.metadata.create_all(engine)
    db = SessionLocal()

    existing = db.query(InterviewQuestion).filter(
        InterviewQuestion.source == "xlsx_import"
    ).count()
    print(f"Existing imported questions: {existing}")

    if existing > 0 and not dry_run:
        print("Clearing previous xlsx imports...")
        db.query(InterviewQuestion).filter(
            InterviewQuestion.source == "xlsx_import"
        ).delete()
        db.commit()

    total_imported = 0
    files = [f for f in os.listdir(questions_dir) if f.endswith(".xlsx")]
    print(f"Found {len(files)} xlsx files")

    for filename in sorted(files):
        cat = _detect_category(filename)
        if not cat:
            print(f"SKIP: {filename!r} — cannot detect category")
            continue
        skill_tag, category, node_id = cat

        path = os.path.join(questions_dir, filename)
        wb = openpyxl.load_workbook(path)
        ws = wb[wb.sheetnames[0]]

        questions = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[1]:
                questions.append({
                    "question": str(row[1]).strip(),
                    "answer": str(row[2]).strip() if row[2] else None,
                })

        print(f"\n{filename!r}: {len(questions)} questions -> {skill_tag}")

        for i, q in enumerate(questions):
            difficulty = assign_difficulty(i, len(questions))
            record = InterviewQuestion(
                node_id=None,  # FK constraint — set via graph mapping later
                skill_tag=skill_tag,
                question=q["question"],
                question_type="technical",
                question_category=category,
                difficulty=difficulty,
                source="xlsx_import",
                answer_key=q["answer"],
            )
            if not dry_run:
                db.add(record)
            total_imported += 1

    if not dry_run:
        db.commit()
        count = db.query(InterviewQuestion).filter(
            InterviewQuestion.source == "xlsx_import"
        ).count()
        print(f"\nDone: {count} questions imported.")
    else:
        print(f"\n[DRY RUN] Would import {total_imported} questions.")

    db.close()


if __name__ == "__main__":
    dry_run = "--dry-run" in sys.argv
    args = [a for a in sys.argv[1:] if not a.startswith("--")]
    questions_dir = args[0] if args else r"C:\Users\liu\Desktop\面试题目"
    import_from_dir(questions_dir, dry_run=dry_run)
